"""
datasheet_importer.py
---------------------

Minimal datasheet importer using only Python standard library.
Supports CSV and TXT files - no external dependencies required.

Usage example:
    from core import datasheet_importer
    tables = datasheet_importer.ingest_file("inventory.csv")
    for name, data in tables:
        print("Table:", name)
        print("Columns:", data['columns'])
        print("First row:", data['rows'][0] if data['rows'] else None)

Supported formats:
- CSV files
- TXT files (tab/comma separated or raw text)

Note: Excel support requires openpyxl, which can be installed optionally.
"""

import csv
import re
from pathlib import Path
from typing import List, Dict, Any, Tuple

# ----------------------------------------------------------------------
# Data structure and utility helpers
# ----------------------------------------------------------------------

def slugify(name: str) -> str:
    """
    Turn an arbitrary string into a safe SQLite table name.
    
    Examples:
        "Product Specs 2025!" → "product_specs_2025"
        "Voltage (V)"         → "voltage_v"
    """
    if not name:
        return "table"
    
    name = str(name).strip().lower()
    name = re.sub(r"\s+", "_", name)  # collapse spaces → "_"
    name = re.sub(r"[^\w\d_]", "", name)  # strip symbols
    return name or "table"


def normalize_columns(columns: List[str]) -> List[str]:
    """
    Normalize column names so they are safe to use in SQLite.
    
    Args:
        columns: List of column names
        
    Returns:
        List of normalized column names
    """
    normalized = []
    for i, col in enumerate(columns, start=1):
        clean_name = slugify(str(col)) if col else f"col_{i}"
        normalized.append(clean_name)
    return normalized


def detect_delimiter(sample_text: str) -> str:
    """
    Detect the most likely delimiter in a text sample.
    
    Args:
        sample_text: Sample of the file content
        
    Returns:
        Most likely delimiter character
    """
    # Count potential delimiters in first few lines
    lines = sample_text.split('\n')[:5]  # Check first 5 lines
    sample = '\n'.join(lines)
    
    tab_count = sample.count('\t')
    comma_count = sample.count(',')
    semicolon_count = sample.count(';')
    pipe_count = sample.count('|')
    
    # Return delimiter with highest count
    delimiters = [('\t', tab_count), (',', comma_count), (';', semicolon_count), ('|', pipe_count)]
    delimiter, _ = max(delimiters, key=lambda x: x[1])
    
    return delimiter if _ > 0 else ','  # Default to comma


# ----------------------------------------------------------------------
# Parsers for different file formats
# ----------------------------------------------------------------------

def parse_csv(path: Path) -> List[Tuple[str, Dict[str, Any]]]:
    """
    Parse a CSV or TSV file using Python's built-in csv module.
    
    Returns:
        List of (table_name, data_dict) tuples where data_dict contains:
        - 'columns': list of column names
        - 'rows': list of lists (each inner list is a row)
    """
    try:
        # Read sample to detect delimiter
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            sample = f.read(1024)
            delimiter = detect_delimiter(sample)
        
        # Read the full file
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            # Try to detect if first line is header
            reader = csv.reader(f, delimiter=delimiter)
            rows = list(reader)
        
        if not rows:
            return [(path.stem, {'columns': ['text'], 'rows': []})]
        
        # Use first row as headers if it looks like headers
        # (contains non-numeric values or is different from other rows)
        if len(rows) > 1:
            first_row = rows[0]
            second_row = rows[1] if len(rows) > 1 else []
            
            # Check if first row looks like headers
            is_header = any(
                not _is_numeric(cell) for cell in first_row
            ) or len(first_row) != len(second_row)
            
            if is_header:
                columns = normalize_columns(first_row)
                data_rows = rows[1:]
            else:
                columns = normalize_columns([f"col_{i}" for i in range(1, len(first_row) + 1)])
                data_rows = rows
        else:
            columns = normalize_columns([f"col_{i}" for i in range(1, len(rows[0]) + 1)])
            data_rows = rows
        
        return [(path.stem, {'columns': columns, 'rows': data_rows})]
        
    except Exception as e:
        # Fallback: treat as plain text
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            return [(path.stem, {'columns': ['text'], 'rows': [[line] for line in lines]})]
        except:
            return [(path.stem, {'columns': ['text'], 'rows': []})]


def parse_excel(path: Path) -> List[Tuple[str, Dict[str, Any]]]:
    """
    Parse an Excel file. Requires openpyxl for .xlsx or xlrd for .xls files.
    
    Returns:
        List of (table_name, data_dict) tuples for each sheet
    """
    try:
        # Try to import openpyxl for modern Excel files
        if path.suffix.lower() == '.xlsx':
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(path, data_only=True)
                results = []
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Get all rows with data
                    rows = []
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            rows.append([str(cell) if cell is not None else '' for cell in row])
                    
                    if rows:
                        # Use first row as headers if it exists
                        if len(rows) > 1:
                            columns = normalize_columns(rows[0])
                            data_rows = rows[1:]
                        else:
                            columns = normalize_columns([f"col_{i}" for i in range(1, len(rows[0]) + 1)])
                            data_rows = rows
                        
                        table_name = f"{path.stem}_{sheet_name}" if len(workbook.sheetnames) > 1 else path.stem
                        results.append((table_name, {'columns': columns, 'rows': data_rows}))
                
                return results
                
            except ImportError:
                raise ImportError("openpyxl is required to read .xlsx files. Install it with: pip install openpyxl")
        
        else:  # .xls files
            try:
                import xlrd
                workbook = xlrd.open_workbook(path)
                results = []
                
                for sheet_name in workbook.sheet_names():
                    sheet = workbook.sheet_by_name(sheet_name)
                    
                    rows = []
                    for row_idx in range(sheet.nrows):
                        row = [str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
                        rows.append(row)
                    
                    if rows:
                        if len(rows) > 1:
                            columns = normalize_columns(rows[0])
                            data_rows = rows[1:]
                        else:
                            columns = normalize_columns([f"col_{i}" for i in range(1, len(rows[0]) + 1)])
                            data_rows = rows
                        
                        table_name = f"{path.stem}_{sheet_name}" if len(workbook.sheet_names()) > 1 else path.stem
                        results.append((table_name, {'columns': columns, 'rows': data_rows}))
                
                return results
                
            except ImportError:
                raise ImportError("xlrd is required to read .xls files. Install it with: pip install xlrd")
                
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {e}")


def parse_txt(path: Path) -> List[Tuple[str, Dict[str, Any]]]:
    """
    Parse a plain-text file.
    
    Attempts to detect delimiter and parse as structured data.
    Falls back to treating each line as a text row if parsing fails.
    """
    try:
        text = path.read_text(encoding="utf-8", errors="ignore")
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        
        if not lines:
            return [(path.stem, {'columns': ['text'], 'rows': []})]
        
        # Try to detect if it's structured data
        delimiter = detect_delimiter('\n'.join(lines[:5]))
        
        # Check if multiple lines have the same number of fields with this delimiter
        field_counts = [len(line.split(delimiter)) for line in lines[:10]]
        if len(set(field_counts)) == 1 and field_counts[0] > 1:
            # Looks like structured data
            rows = [line.split(delimiter) for line in lines]
            
            # Use first row as headers if it looks like headers
            if len(rows) > 1 and not _is_numeric(rows[0][0]):
                columns = normalize_columns(rows[0])
                data_rows = rows[1:]
            else:
                columns = normalize_columns([f"col_{i}" for i in range(1, len(rows[0]) + 1)])
                data_rows = rows
            
            return [(path.stem, {'columns': columns, 'rows': data_rows})]
        else:
            # Treat as plain text
            return [(path.stem, {'columns': ['text'], 'rows': [[line] for line in lines]})]
            
    except Exception:
        # Final fallback
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            return [(path.stem, {'columns': ['text'], 'rows': [[line] for line in lines]})]
        except:
            return [(path.stem, {'columns': ['text'], 'rows': []})]


def _is_numeric(value: str) -> bool:
    """Check if a string represents a number."""
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


# ----------------------------------------------------------------------
# Dispatcher
# ----------------------------------------------------------------------

PARSERS = {
    ".csv": parse_csv,
    ".tsv": parse_csv,
    ".txt": parse_txt,
}

# Excel parsers are optional - only add if dependencies are available
try:
    import openpyxl
    PARSERS[".xlsx"] = parse_excel
except ImportError:
    pass

try:
    import xlrd
    PARSERS[".xls"] = parse_excel
except ImportError:
    pass


def ingest_file(path: str) -> List[Tuple[str, Dict[str, Any]]]:
    """
    Ingest a file into structured data.

    Args:
        path: Path to file (CSV, TXT, or Excel if libraries available)

    Returns:
        List of (table_name, data_dict) tuples where data_dict contains:
        - 'columns': list of column names
        - 'rows': list of lists (each inner list represents a row)

    Raises:
        FileNotFoundError: if path does not exist
        ValueError: if file extension is unsupported
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"File not found: {path}")
    
    ext = p.suffix.lower()
    parser = PARSERS.get(ext)
    
    if not parser:
        supported_formats = list(PARSERS.keys())
        raise ValueError(f"Unsupported file type: {ext}. Supported formats: {supported_formats}")
    
    return parser(p)


def get_supported_extensions() -> List[str]:
    """
    Get list of supported file extensions.
    
    Returns:
        List of supported file extensions (with dots)
    """
    return list(PARSERS.keys())
